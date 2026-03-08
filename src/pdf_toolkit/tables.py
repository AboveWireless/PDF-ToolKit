from __future__ import annotations

from pathlib import Path
import csv

import openpyxl
import pdfplumber

from pdf_toolkit.core import ensure_dir, ensure_parent_dir, parse_page_spec
from pdf_toolkit.ocr import run_ocr
from pdf_toolkit.reporting import write_json


def extract_tables_to_files(
    input_path: Path,
    output_dir: Path,
    *,
    page_spec: str | None,
    format_name: str,
    ocr_first: bool,
    ocr_language: str,
    temp_dir: Path,
) -> dict[str, object]:
    ensure_dir(output_dir)
    source_path = input_path
    if ocr_first:
        temp_ocr_output = temp_dir / "ocr" / f"{input_path.stem}-tables-ocr.pdf"
        run_ocr(
            input_path,
            temp_ocr_output,
            language=ocr_language,
            skip_existing_text=True,
            text_output=None,
            json_output=None,
            force=False,
            temp_dir=temp_dir,
        )
        source_path = temp_ocr_output

    csv_outputs: list[Path] = []
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    manifest: dict[str, object] = {
        "input_path": str(input_path),
        "source_path": str(source_path),
        "tables": [],
    }

    with pdfplumber.open(str(source_path)) as pdf:
        selected = set(parse_page_spec(page_spec, len(pdf.pages), allow_duplicates=False)) if page_spec else None
        for page_index, page in enumerate(pdf.pages, start=1):
            if selected is not None and page_index - 1 not in selected:
                continue
            tables = page.extract_tables()
            for table_index, rows in enumerate(tables, start=1):
                clean_rows = [[cell or "" for cell in row] for row in rows]
                table_name = f"page-{page_index:03d}-table-{table_index:02d}"
                entry: dict[str, object] = {
                    "page": page_index,
                    "table_index": table_index,
                    "row_count": len(clean_rows),
                }
                if format_name in {"csv", "all"}:
                    csv_path = output_dir / f"{input_path.stem}-{table_name}.csv"
                    with csv_path.open("w", newline="", encoding="utf-8") as handle:
                        writer = csv.writer(handle)
                        writer.writerows(clean_rows)
                    csv_outputs.append(csv_path)
                    entry["csv_path"] = str(csv_path)
                if format_name in {"xlsx", "all"}:
                    sheet = workbook.create_sheet(title=f"p{page_index}_t{table_index}")
                    for row in clean_rows:
                        sheet.append(row)
                manifest["tables"].append(entry)

    outputs: list[Path] = list(csv_outputs)
    if format_name in {"xlsx", "all"} and workbook.sheetnames:
        xlsx_path = output_dir / f"{input_path.stem}-tables.xlsx"
        ensure_parent_dir(xlsx_path)
        workbook.save(xlsx_path)
        outputs.append(xlsx_path)
        manifest["xlsx_path"] = str(xlsx_path)
    if format_name in {"json", "all"}:
        json_path = output_dir / f"{input_path.stem}-tables.json"
        write_json(manifest, json_path)
        outputs.append(json_path)

    return {
        "outputs": outputs,
        "details": manifest,
    }
